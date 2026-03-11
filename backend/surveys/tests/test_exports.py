from pathlib import Path
from tempfile import TemporaryDirectory
from zipfile import ZipFile
from importlib.util import find_spec

from django.conf import settings
from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.utils import timezone
from rest_framework import status
from rest_framework.test import APIClient

from surveys.models import Answer, Choice, ExportJob, Page, Question, Survey, SurveyResponse
from surveys.services.exports.common import format_answer_for_export

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional test dependency
    load_workbook = None

WEASYPRINT_AVAILABLE = find_spec("weasyprint") is not None
PPTX_AVAILABLE = find_spec("pptx") is not None

User = get_user_model()


class ExportJobApiTests(TestCase):
    def setUp(self):
        self.temp_dir = TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.settings_override = override_settings(
            CELERY_TASK_ALWAYS_EAGER=True,
            CELERY_TASK_EAGER_PROPAGATES=True,
            MEDIA_ROOT=self.temp_dir.name,
            API_BASE_URL="http://testserver/api",
        )
        self.settings_override.enable()
        self.addCleanup(self.settings_override.disable)

        self.user = User.objects.create_user(
            username="export-owner",
            email="export-owner@example.com",
            password="TestPass123!",
        )
        self.client = APIClient()
        self.client.force_authenticate(self.user)

        self.survey = Survey.objects.create(
            user=self.user,
            title="Export Survey",
            description="Survey used for export tests.",
            status=Survey.Status.ACTIVE,
            theme={"primary_color": "#f97316"},
        )
        self.page = Page.objects.create(survey=self.survey, title="Page 1", order=1)
        self.choice_question = Question.objects.create(
            page=self.page,
            question_type=Question.QuestionType.MULTIPLE_CHOICE_SINGLE,
            text="How satisfied are you?",
            order=1,
        )
        self.choice_yes = Choice.objects.create(
            question=self.choice_question,
            text="Great",
            order=1,
        )
        self.choice_no = Choice.objects.create(
            question=self.choice_question,
            text="Okay",
            order=2,
        )
        self.text_question = Question.objects.create(
            page=self.page,
            question_type=Question.QuestionType.SHORT_TEXT,
            text="Why did you choose that score?",
            order=2,
        )

        self._create_response(self.choice_yes.id, "Fast delivery and clear communication.")
        self._create_response(self.choice_no.id, "It was acceptable overall.")

    def _create_response(self, choice_id, text_value):
        response = SurveyResponse.objects.create(
            survey=self.survey,
            status=SurveyResponse.Status.COMPLETED,
            current_page=self.page,
            respondent_email=f"user-{SurveyResponse.objects.count()}@example.com",
        )
        completed_at = timezone.now()
        SurveyResponse.objects.filter(id=response.id).update(
            completed_at=completed_at,
            duration_seconds=42,
        )
        response.refresh_from_db()

        Answer.objects.create(
            response=response,
            question=self.choice_question,
            choice_ids=[str(choice_id)],
        )
        Answer.objects.create(
            response=response,
            question=self.text_question,
            text_value=text_value,
        )
        return response

    def _create_export(self, export_format):
        response = self.client.post(
            f"/api/surveys/{self.survey.id}/exports/",
            {
                "format": export_format,
                "config": {
                    "filters": {"status": "completed"},
                    "include_individual_responses": True,
                },
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(response.data["format"], export_format)
        self.assertEqual(response.data["status"], ExportJob.Status.COMPLETED)
        self.assertTrue(response.data["file_url"])
        return response.data

    def test_pdf_export_job_generates_pdf_file(self):
        if not WEASYPRINT_AVAILABLE:
            self.skipTest("weasyprint is not installed in this environment")

        job = self._create_export("pdf")
        export_job = ExportJob.objects.get(id=job["id"])
        file_path = Path(settings.MEDIA_ROOT) / "exports" / str(self.survey.id) / f"{export_job.id}.pdf"

        self.assertTrue(file_path.exists())
        self.assertTrue(file_path.read_bytes().startswith(b"%PDF"))

        detail = self.client.get(f"/api/surveys/{self.survey.id}/exports/{export_job.id}/")
        self.assertEqual(detail.status_code, status.HTTP_200_OK)
        self.assertEqual(detail.data["status"], ExportJob.Status.COMPLETED)

    def test_xlsx_export_job_generates_workbook_file(self):
        if load_workbook is None:
            self.skipTest("openpyxl is not installed in this environment")

        job = self._create_export("xlsx")
        export_job = ExportJob.objects.get(id=job["id"])
        file_path = Path(settings.MEDIA_ROOT) / "exports" / str(self.survey.id) / f"{export_job.id}.xlsx"

        self.assertTrue(file_path.exists())
        with ZipFile(file_path) as archive:
            self.assertIn("xl/workbook.xml", archive.namelist())
        workbook = load_workbook(file_path)
        self.assertIn("Summary", workbook.sheetnames)
        self.assertIn("Raw Responses", workbook.sheetnames)
        self.assertIn("Raw Questions", workbook.sheetnames)
        self.assertEqual(workbook["Raw Questions"]["A2"].value, "Status")

    def test_pptx_export_job_generates_presentation_file(self):
        if not PPTX_AVAILABLE:
            self.skipTest("python-pptx is not installed in this environment")

        job = self._create_export("pptx")
        export_job = ExportJob.objects.get(id=job["id"])
        file_path = Path(settings.MEDIA_ROOT) / "exports" / str(self.survey.id) / f"{export_job.id}.pptx"

        self.assertTrue(file_path.exists())
        with ZipFile(file_path) as archive:
            self.assertIn("ppt/presentation.xml", archive.namelist())


class ExportAnswerFormattingTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(
            username="export-format-owner",
            email="export-format-owner@example.com",
            password="TestPass123!",
        )
        self.survey = Survey.objects.create(user=self.user, title="Format Survey")
        self.page = Page.objects.create(survey=self.survey, title="Page 1", order=1)

    def test_format_answer_for_export_handles_open_ended_rows(self):
        question = Question.objects.create(
            page=self.page,
            question_type=Question.QuestionType.OPEN_ENDED,
            text="Bandwidth snapshot",
            order=1,
            settings={
                "rows": ["Commodity", "NIX"],
                "allow_other": True,
            },
        )
        response = SurveyResponse.objects.create(
            survey=self.survey,
            status=SurveyResponse.Status.COMPLETED,
            current_page=self.page,
        )
        answer = Answer.objects.create(
            response=response,
            question=question,
            matrix_data={
                "Commodity": "100 Mbps",
                "NIX": "50 Mbps",
                "__other__": "Cache",
            },
        )

        self.assertEqual(
            format_answer_for_export(answer),
            "Commodity: 100 Mbps; NIX: 50 Mbps; Other: Cache",
        )

    def test_format_answer_for_export_handles_matrix_plus_cells(self):
        question = Question.objects.create(
            page=self.page,
            question_type=Question.QuestionType.MATRIX_PLUS,
            text="List",
            order=1,
            settings={
                "rows": ["Item 1", "Item 2"],
                "columns": ["Col1", "Col2"],
                "dropdown_options": ["Ddown1", "Ddown2", "Ddown3"],
            },
        )
        response = SurveyResponse.objects.create(
            survey=self.survey,
            status=SurveyResponse.Status.COMPLETED,
            current_page=self.page,
        )
        answer = Answer.objects.create(
            response=response,
            question=question,
            matrix_data={
                "Item 1": {"Col1": "Ddown1", "Col2": "Ddown2"},
                "Item 2": {"Col1": "Ddown3", "Col2": "Ddown1"},
            },
        )

        self.assertEqual(
            format_answer_for_export(answer),
            "Item 1 / Col1: Ddown1; Item 1 / Col2: Ddown2; Item 2 / Col1: Ddown3; Item 2 / Col2: Ddown1",
        )
