from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _normalize_hex_color(value):
    hex_value = (value or "#2563eb").strip().lstrip("#")
    if len(hex_value) == 3:
        hex_value = "".join(character * 2 for character in hex_value)
    return hex_value[:6].upper()


class XLSXExportService:
    def generate(self, survey, analytics_data, config) -> bytes:
        workbook = Workbook()
        brand_color = _normalize_hex_color(analytics_data["branding"]["color"])

        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        self._build_summary_sheet(summary_sheet, survey, analytics_data, brand_color)

        raw_sheet = workbook.create_sheet("Raw Responses")
        self._build_raw_sheet(raw_sheet, analytics_data, brand_color)

        transposed_raw_sheet = workbook.create_sheet("Raw Questions")
        self._build_transposed_raw_sheet(transposed_raw_sheet, analytics_data, brand_color)

        for index, cross_tab in enumerate(analytics_data["cross_tabs"], start=1):
            cross_tab_sheet = workbook.create_sheet(f"Cross Tab {index}")
            self._build_cross_tab_sheet(cross_tab_sheet, cross_tab, brand_color)

        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def _build_summary_sheet(self, worksheet, survey, analytics_data, brand_color):
        summary = analytics_data["summary"]
        fill = PatternFill("solid", fgColor=brand_color)
        header_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        border = Border(
            left=Side(style="thin", color="D9E2EC"),
            right=Side(style="thin", color="D9E2EC"),
            top=Side(style="thin", color="D9E2EC"),
            bottom=Side(style="thin", color="D9E2EC"),
        )

        worksheet["A1"] = survey.title
        worksheet["A1"].font = Font(size=16, bold=True)
        worksheet["A3"] = "Generated"
        worksheet["B3"] = analytics_data["generated_at"].strftime("%Y-%m-%d %H:%M UTC")
        worksheet["A4"] = "Survey slug"
        worksheet["B4"] = survey.slug

        metric_rows = [
            ("Total Responses", summary["total_responses"]),
            ("Completion Rate", summary["completion_rate"] / 100),
            ("Average Duration (sec)", summary["average_duration_seconds"]),
            ("Responses In Progress", summary["in_progress_count"]),
        ]
        start_row = 6
        worksheet[f"A{start_row}"] = "Metric"
        worksheet[f"B{start_row}"] = "Value"
        for cell in (worksheet[f"A{start_row}"], worksheet[f"B{start_row}"]):
            cell.fill = fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_offset, (label, value) in enumerate(metric_rows, start=1):
            worksheet.cell(start_row + row_offset, 1, label)
            value_cell = worksheet.cell(start_row + row_offset, 2, value)
            if "Rate" in label:
                value_cell.number_format = "0.0%"

        current_row = start_row + len(metric_rows) + 3
        for section in analytics_data["questions"]:
            worksheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            header_cell = worksheet.cell(current_row, 1, section["text"])
            header_cell.fill = fill
            header_cell.font = header_font
            header_cell.alignment = Alignment(vertical="center")
            current_row += 1

            for column_index, title in enumerate(section["table_columns"], start=1):
                cell = worksheet.cell(current_row, column_index, title)
                cell.font = bold_font
                cell.border = border
                cell.fill = PatternFill("solid", fgColor="EEF4FF")
            current_row += 1

            if section["analytics_type"] == "matrix":
                for row in section["table_rows"]:
                    worksheet.cell(current_row, 1, row["label"])
                    for value_index, value in enumerate(row["values"], start=2):
                        worksheet.cell(current_row, value_index, value)
                    current_row += 1
            elif section["analytics_type"] == "demographics":
                for field_section in section["field_sections"]:
                    worksheet.cell(current_row, 1, field_section["field_name"]).font = bold_font
                    current_row += 1
                    for row in field_section["rows"]:
                        worksheet.cell(current_row, 1, row["label"])
                        worksheet.cell(current_row, 2, row["count"])
                        percentage_cell = worksheet.cell(current_row, 3, row["percentage"] / 100)
                        percentage_cell.number_format = "0.0%"
                        current_row += 1
            else:
                data_start_row = current_row
                for row in section["table_rows"]:
                    worksheet.cell(current_row, 1, row["label"])
                    worksheet.cell(current_row, 2, row["count"])
                    percentage_value = row.get("percentage")
                    if percentage_value not in ("", None):
                        percentage_cell = worksheet.cell(current_row, 3, percentage_value / 100)
                        percentage_cell.number_format = "0.0%"
                    current_row += 1

                if section["table_rows"]:
                    worksheet.conditional_formatting.add(
                        f"B{data_start_row}:B{current_row - 1}",
                        DataBarRule(
                            start_type="num",
                            start_value=0,
                            end_type="max",
                            color=brand_color,
                        ),
                    )

            current_row += 2

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = border
        self._autofit_columns(worksheet)

    def _build_raw_sheet(self, worksheet, analytics_data, brand_color):
        questions = analytics_data["raw_questions"]
        header_font = Font(color="FFFFFF", bold=True)
        fill = PatternFill("solid", fgColor=brand_color)
        headers = [
            "Response ID",
            "Status",
            "Started",
            "Completed",
            "Duration (sec)",
            "Collector",
            "Email",
            "IP Address",
            *[question.text for question in questions],
        ]
        for index, title in enumerate(headers, start=1):
            cell = worksheet.cell(1, index, title)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        for row_index, row in enumerate(analytics_data["raw_responses"], start=2):
            values = [
                row["id"],
                row["status"],
                row["started_at"].strftime("%Y-%m-%d %H:%M:%S") if row["started_at"] else "",
                row["completed_at"].strftime("%Y-%m-%d %H:%M:%S") if row["completed_at"] else "",
                row["duration_seconds"] or "",
                row["collector_name"],
                row["respondent_email"],
                row["ip_address"],
                *list(row["values"].values()),
            ]
            for column_index, value in enumerate(values, start=1):
                worksheet.cell(row_index, column_index, value)

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        self._autofit_columns(worksheet)

    def _build_transposed_raw_sheet(self, worksheet, analytics_data, brand_color):
        questions = analytics_data["raw_questions"]
        responses = analytics_data["raw_responses"]
        header_font = Font(color="FFFFFF", bold=True)
        fill = PatternFill("solid", fgColor=brand_color)

        headers = ["Field", *[row["id"] for row in responses]]
        for index, title in enumerate(headers, start=1):
            cell = worksheet.cell(1, index, title)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")

        metadata_rows = [
            ("Status", [row["status"] for row in responses]),
            (
                "Started",
                [
                    row["started_at"].strftime("%Y-%m-%d %H:%M:%S")
                    if row["started_at"]
                    else ""
                    for row in responses
                ],
            ),
            (
                "Completed",
                [
                    row["completed_at"].strftime("%Y-%m-%d %H:%M:%S")
                    if row["completed_at"]
                    else ""
                    for row in responses
                ],
            ),
            ("Duration (sec)", [row["duration_seconds"] or "" for row in responses]),
            ("Collector", [row["collector_name"] for row in responses]),
            ("Email", [row["respondent_email"] for row in responses]),
            ("IP Address", [row["ip_address"] for row in responses]),
        ]

        question_rows = [
            (
                question.text,
                [row["values"].get(str(question.id), "") for row in responses],
            )
            for question in questions
        ]

        for row_index, (label, values) in enumerate([*metadata_rows, *question_rows], start=2):
            worksheet.cell(row_index, 1, label)
            for column_index, value in enumerate(values, start=2):
                worksheet.cell(row_index, column_index, value)

        worksheet.freeze_panes = "B2"
        worksheet.auto_filter.ref = worksheet.dimensions
        self._autofit_columns(worksheet)

    def _build_cross_tab_sheet(self, worksheet, cross_tab, brand_color):
        fill = PatternFill("solid", fgColor=brand_color)
        header_font = Font(color="FFFFFF", bold=True)

        worksheet["A1"] = cross_tab["title"]
        worksheet["A1"].font = Font(size=14, bold=True)

        headers = ["Row label", *[item["col_label"] for item in cross_tab["col_totals"]], "Total"]
        for index, title in enumerate(headers, start=1):
            cell = worksheet.cell(3, index, title)
            cell.fill = fill
            cell.font = header_font

        for row_index, row in enumerate(cross_tab["matrix"], start=4):
            worksheet.cell(row_index, 1, row["row_label"])
            for column_index, cell_data in enumerate(row["cells"], start=2):
                worksheet.cell(row_index, column_index, cell_data["count"])
            worksheet.cell(row_index, len(headers), row["row_total"])

        totals_row = len(cross_tab["matrix"]) + 4
        worksheet.cell(totals_row, 1, "Column total").font = Font(bold=True)
        for column_index, column_data in enumerate(cross_tab["col_totals"], start=2):
            worksheet.cell(totals_row, column_index, column_data["total"]).font = Font(bold=True)
        worksheet.cell(totals_row, len(headers), cross_tab["grand_total"]).font = Font(bold=True)

        worksheet.cell(totals_row + 2, 1, "Chi-square statistic")
        worksheet.cell(totals_row + 2, 2, cross_tab["chi_square"]["statistic"])
        worksheet.cell(totals_row + 3, 1, "p-value")
        worksheet.cell(totals_row + 3, 2, cross_tab["chi_square"]["p_value"])

        if cross_tab["matrix"]:
            worksheet.conditional_formatting.add(
                f"B4:{get_column_letter(len(headers) - 1)}{totals_row - 1}",
                ColorScaleRule(
                    start_type="min",
                    start_color="E8F5E9",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFF9C4",
                    end_type="max",
                    end_color="FFCDD2",
                ),
            )
        self._autofit_columns(worksheet)

    def _autofit_columns(self, worksheet):
        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max_length + 3, 42)
